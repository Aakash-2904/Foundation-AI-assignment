import os, re, sys, time, json, argparse, textwrap
from pathlib import Path
import openpyxl
import ollama
try:
    import numpy as np
    from sentence_transformers import SentenceTransformer
    from sklearn.metrics.pairwise import cosine_similarity as _cos_sim

    # we are using all-MiniLM-L6-v2 embedding model which is 80 MB cause it is fast nd accurate for semantic similarity.
    _EMBED_MODEL = SentenceTransformer("all-MiniLM-L6-v2")
    COSINE_AVAILABLE = True
    print("[Cosine] sentence-transformers loaded — cosine similarity enabled.")
except ImportError:
    COSINE_AVAILABLE = False
    print("[Cosine] sentence-transformers not found — cosine similarity will be N/A.")
    print("         Install with:  pip install sentence-transformers scikit-learn numpy")

# CONFIG
DEFAULT_EXCEL  = "jobs_output.xlsx"       # Excel file scraped by the job scraper
OUTPUT_DIR     = "tailored_resumes_ollama_5"  # where all the .tex files will land
DEFAULT_MAX    = 20                        # cap on how many jobs to process in one run
DEFAULT_MODEL  = "llama3.2"               # Ollama model to use (swap for mistral, phi3, etc.)

COL_TITLE = "Position Title"
COL_COMPANY = "Company"
COL_JD = "Full Job Description"
COL_TAGS = "Tags / Skills"
COL_QUALS = "Qualifications"
COL_RESP = "Responsibilities"

# PROMPTS
# The system prompt is the "instruction manual" we hand to the model.
# It defines exactly what JSON shape we expect back, and lays down
# strict rules so the model doesn't hallucinate metrics or drop skills.
SYSTEM_META = """\
You are an expert ATS resume tailor. Return your response as plain KEY: VALUE lines — no JSON, no markdown, no brackets, no quotes, no explanation.

FORMAT RULES (read carefully):
- Every line must be exactly:   KEY: value
- Bullets use a pipe separator:  exp1_bullets: did X achieving Y% | did Z saving $N | improved W by N%
- Skills use comma separation:   skill_items_1: Python, TensorFlow, SQL
- One line per key — never break a value across multiple lines
- No blank lines between keys
- No colons inside values (use a dash instead)

OUTPUT ALL OF THESE KEYS IN ORDER:
name
contact
summary
skill_cat_1
skill_items_1
skill_cat_2
skill_items_2
skill_cat_3
skill_items_3
skill_cat_4
skill_items_4
exp1_title
exp1_company
exp1_location
exp1_dates
exp1_bullets
exp2_title
exp2_company
exp2_location
exp2_dates
exp2_bullets
edu1_degree
edu1_institution
edu1_location
edu1_dates
edu1_detail
edu2_degree
edu2_institution
edu2_location
edu2_dates
edu2_detail
proj1_name
proj1_bullets
proj2_name
proj2_bullets
proj3_name
proj3_bullets
certifications
match_score
keywords

CONTENT RULES:
1. Keep ALL quantified metrics from the base resume. Never drop a percentage or number.
2. Rephrase bullets to embed keywords from the job description naturally.
3. Do NOT invent jobs, companies, degrees, or metrics not in the base resume.
4. summary must contain the exact job title from the JD.
5. match_score is an integer 0-100 estimating how well the resume fits the JD.
6. keywords is a comma-separated list of 5 matched JD keywords.
"""


def _jd(job: dict, limit: int = 2000) -> str:

    return "\n".join(filter(None, [
        str(job.get(COL_JD,    "") or ""),
        str(job.get(COL_QUALS, "") or ""),
        str(job.get(COL_RESP,  "") or ""),
        str(job.get(COL_TAGS,  "") or ""),
    ]))[:limit]


def meta_prompt(base_resume: str, job: dict) -> str:
    return (
        f"JOB TITLE : {job.get(COL_TITLE,  'N/A')}\n"
        f"COMPANY   : {job.get(COL_COMPANY, 'N/A')}\n\n"
        f"JOB DESCRIPTION:\n{_jd(job)}\n\n"
        f"BASE RESUME:\n{base_resume}"
    )

# lat buildr


def esc(s: str) -> str:
    """
    Escape characters that would break LaTeX compilation.
    Order matters: backslash must be replaced first, otherwise
    the replacement strings themselves get double-escaped.
    """
    s = str(s)
    replacements = [
        ("\\", "\\textbackslash{}"),
        ("&",  "\\&"),
        ("%",  "\\%"),
        ("$",  "\\$"),
        ("#",  "\\#"),
        ("_",  "\\_"),
        ("{",  "\\{"),
        ("}",  "\\}"),
        ("~",  "\\textasciitilde{}"),
        ("^",  "\\textasciicircum{}"),
    ]
    for char, repl in replacements:
        s = s.replace(char, repl)
    return s


def build_latex(d: dict) -> str:

    def section(title: str) -> str:
        return f"\n\\section*{{{title}}}\n"

    def split_pipes(s: str) -> list[str]:

        if not s or not str(s).strip():
            return []
        return [p.strip() for p in str(s).split("|") if p.strip()]

    def itemize(bullets: list) -> str:
        if not bullets:
            return ""
        items = "\n".join(f"  \\item {esc(b)}" for b in bullets)
        return f"\\begin{{itemize}}\n{items}\n\\end{{itemize}}\n"

    lines = []

    lines.append(r"\documentclass[10pt]{article}")
    lines.append(r"\usepackage[top=0.5in, bottom=0.5in, left=0.65in, right=0.65in]{geometry}")
    lines.append(r"\usepackage[T1]{fontenc}")
    lines.append(r"\usepackage{enumitem}")
    lines.append(r"\usepackage[hidelinks]{hyperref}")
    lines.append(r"\usepackage{titlesec}")
    lines.append(r"\titleformat{\section}{\large\bfseries}{}{0em}{}[\hrule]")
    lines.append(r"\titlespacing*{\section}{0pt}{8pt}{4pt}")
    lines.append(r"\setlist[itemize]{noitemsep, topsep=2pt, leftmargin=1.5em}")
    lines.append(r"\pagestyle{empty}")
    lines.append(r"\begin{document}")
    lines.append("")

    name    = esc(d.get("name",    "Candidate Name"))
    contact = esc(d.get("contact", ""))
    lines.append(r"\begin{center}")
    lines.append(f"  {{\\LARGE \\textbf{{{name}}}}} \\\\[4pt]")
    lines.append(f"  {contact}")
    lines.append(r"\end{center}")
    lines.append(r"\vspace{4pt}")

    summary = d.get("summary", "")
    if summary:
        lines.append(section("Professional Summary"))
        lines.append(esc(summary))
        lines.append("")
    skill_lines = []
    for n in range(1, 7):
        cat   = d.get(f"skill_cat_{n}",   "")
        items = d.get(f"skill_items_{n}", "")
        if cat and items:
            skill_lines.append(f"  \\item \\textbf{{{esc(cat)}:}} {esc(items)}")
    if skill_lines:
        lines.append(section("Relevant Skills"))
        lines.append(r"\begin{itemize}")
        lines.extend(skill_lines)
        lines.append(r"\end{itemize}")
        lines.append("")
    exp_added = False
    for n in range(1, 4):
        title    = d.get(f"exp{n}_title",    "")
        company  = d.get(f"exp{n}_company",  "")
        location = d.get(f"exp{n}_location", "")
        dates    = d.get(f"exp{n}_dates",    "")
        bullets  = split_pipes(d.get(f"exp{n}_bullets", ""))
        if not title:
            continue
        if not exp_added:
            lines.append(section("Professional Work Experience"))
            exp_added = True
        lines.append(f"\\textbf{{{esc(title)}}} \\hfill \\textit{{{esc(dates)}}}")
        lines.append(f"\\textit{{{esc(company)}, {esc(location)}}}")
        lines.append(itemize(bullets))
    edu_added = False
    for n in range(1, 3):
        degree = d.get(f"edu{n}_degree",      "")
        inst   = d.get(f"edu{n}_institution", "")
        loc    = d.get(f"edu{n}_location",    "")
        dates  = d.get(f"edu{n}_dates",       "")
        detail = d.get(f"edu{n}_detail",      "")
        if not degree:
            continue
        if not edu_added:
            lines.append(section("Education"))
            edu_added = True
        lines.append(f"\\textbf{{{esc(degree)}}} \\hfill \\textit{{{esc(dates)}}}")
        lines.append(f"\\textit{{{esc(inst)}, {esc(loc)}}}")
        if detail:
            lines.append(esc(detail))
        lines.append("")

    proj_added = False
    for n in range(1, 4):
        pname   = d.get(f"proj{n}_name",    "")
        bullets = split_pipes(d.get(f"proj{n}_bullets", ""))
        if not pname:
            continue
        if not proj_added:
            lines.append(section("Projects"))
            proj_added = True
        lines.append(f"\\textbf{{{esc(pname)}}}")
        lines.append(itemize(bullets))

    certs = split_pipes(d.get("certifications", ""))
    if certs:
        lines.append(section("Certifications"))
        lines.append(itemize(certs))

    lines.append(r"\end{document}")
    return "\n".join(lines)


def fallback_latex(base_resume: str, job: dict, reason: str) -> str:
# safely returns 
    def e(s):
        for c in "&%$#_{}~^":
            s = s.replace(c, f"\\{c}")
        return s
    title   = e(str(job.get(COL_TITLE,   "Position")))
    company = e(str(job.get(COL_COMPANY, "Company")))
    return (
        "\\documentclass[11pt]{article}\n\\usepackage[margin=0.75in]{geometry}\n"
        "\\usepackage{parskip}\n\\begin{document}\n\n"
        f"% Tailoring failed: {reason}\n% Job: {title} @ {company}\n\n"
        "\\begin{center}{\\Large \\textbf{" + title + " --- " + company + "}}\\end{center}\n\n"
        + e(base_resume[:3000]) + "\n\n\\end{document}\n"
    )


def load_jobs(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Jobs"]   
    headers = [c.value for c in ws[1]]   # first row = column names
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        job = {headers[i]: (row[i] if i < len(row) and row[i] is not None else "")
               for i in range(len(headers))}
        if job.get(COL_TITLE) and job.get(COL_JD):
            jobs.append(job)
    print(f"Loaded {len(jobs)} jobs from '{path}'")
    return jobs



def safe_filename(idx: int, company: str, title: str) -> str:
    def slug(s):
        # Strip special characters, lowercase, replace spaces with underscores, truncate
        s = re.sub(r"[^\w\s-]", "", str(s).lower())
        return re.sub(r"[\s_-]+", "_", s).strip("_")[:30]
    return f"job_{idx:03d}_{slug(company)}_{slug(title)}.tex"


def compute_cosine_similarity(resume_text: str, jd_text: str) -> float | str:
    if not COSINE_AVAILABLE:
        return "N/A"
    if not resume_text or not jd_text:
        return "N/A"
    try:
        r_text = resume_text[:3000]
        j_text = jd_text[:3000]
        embeddings = _EMBED_MODEL.encode([r_text, j_text], show_progress_bar=False)
        score = float(_cos_sim([embeddings[0]], [embeddings[1]])[0][0])
        return round(score, 4)
    except Exception as e:
        print(f"[Cosine] Error computing similarity: {e}")
        return "N/A"

# ollama

class OllamaResumeAgent:

    def __init__(self, model: str):
        self.model = model
        self._check_ollama()   # fail fast if Ollama isn't running or model isn't pulled

    def _check_ollama(self):
# verification 
        try:
            models    = [m.model for m in ollama.list().models]
            available = [m.split(":")[0] for m in models]
            wanted    = self.model.split(":")[0]
            if wanted not in available:
                print(f"\nModel '{self.model}' not found. Available: {available}")
                print(f"Run:ollama pull {self.model}")
                sys.exit(1)
            print(f"Ollama ready model: {self.model}")
        except Exception as e:
            print(f"\nCannot connect to Ollama: {e}")
            print("Run: ollama serve")
            sys.exit(1)

    def _call(self, system: str, user: str) -> str:
        resp = ollama.chat(
            model=self.model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user",   "content": user},
            ],
            options={"temperature": 0.2, "num_predict": 4096},
        )
        return resp.message.content.strip()

    @staticmethod
    def _parse_kv(raw: str) -> dict:
        KNOWN_KEYS = {
            "name", "contact", "summary",
            "skill_cat_1", "skill_items_1",
            "skill_cat_2", "skill_items_2",
            "skill_cat_3", "skill_items_3",
            "skill_cat_4", "skill_items_4",
            "exp1_title", "exp1_company", "exp1_location",
            "exp1_dates", "exp1_bullets",
            "exp2_title", "exp2_company", "exp2_location",
            "exp2_dates", "exp2_bullets",
            "edu1_degree", "edu1_institution", "edu1_location",
            "edu1_dates",  "edu1_detail",
            "edu2_degree", "edu2_institution", "edu2_location",
            "edu2_dates",  "edu2_detail",
            "proj1_name", "proj1_bullets",
            "proj2_name", "proj2_bullets",
            "proj3_name", "proj3_bullets",
            "certifications", "match_score", "keywords",
        }

        result      = {}
        current_key = None

        for line in raw.splitlines():
            line = line.strip()
            if not line:
                continue

            matched = False
            for key in KNOWN_KEYS:
                if line.lower().startswith(key + ":"):
                    value = line[len(key) + 1:].strip()
                    result[key]  = value
                    current_key  = key
                    matched      = True
                    break

            if not matched and current_key:
                result[current_key] = result[current_key] + " " + line

        try:
            result["match_score"] = int(str(result.get("match_score", 50)).strip())
        except (ValueError, TypeError):
            result["match_score"] = 50

        return result

    def _get_structured_data(self, base_resume: str, job: dict) -> dict:

        prompt = meta_prompt(base_resume, job)
        raw    = self._call(SYSTEM_META, prompt)
        return self._parse_kv(raw)

    def process_job(self, base_resume: str, job: dict,
                    out_dir: Path, idx: int) -> dict:
        title   = str(job.get(COL_TITLE,   "Unknown"))
        company = str(job.get(COL_COMPANY, "Unknown"))
        fpath   = out_dir / safe_filename(idx, company, title)

        print(f"\n  [{idx:>3}]  {title[:45]:<45}  @  {company[:25]}")

        match_score  = "N/A"
        keywords     = []
        cosine_sim   = "N/A"
        resume_text  = ""   # plain-text resume content for similarity computation

        try:
            data        = self._get_structured_data(base_resume, job)
            match_score = data.get("match_score", "N/A")
            keywords    = data.get("keywords",    [])

            latex = build_latex(data)

            summary    = data.get("summary", "")
            skills_txt = " ".join(
                data.get(f"skill_items_{n}", "")
                for n in range(1, 7)
            )
            exp_txt = " ".join(
                data.get(f"exp{n}_bullets", "").replace("|", " ")
                for n in range(1, 4)
            )
            proj_txt = " ".join(
                data.get(f"proj{n}_bullets", "").replace("|", " ")
                for n in range(1, 4)
            )
            resume_text = f"{summary} {skills_txt} {exp_txt} {proj_txt}".strip()

            jd_text    = _jd(job, limit=3000)
            cosine_sim = compute_cosine_similarity(resume_text, jd_text)

            print(f"Data   : score={match_score}  keywords={len(keywords)}  "
                  f"cosine_sim={cosine_sim}")
            print(f"LaTeX  : {len(latex)} chars (Python-generated)")

        except Exception as e:
            latex = fallback_latex(base_resume, job, str(e))
            print(f"{str(e)[:80]} — fallback written")

        fpath.write_text(latex, encoding="utf-8")
        print(f"{fpath}")

        return {
            "index":      idx,
            "file":       str(fpath),
            "title":      title,
            "company":    company,
            "match_score": match_score,
            "keywords":   keywords,
            "cosine_sim": cosine_sim,   
        }

    def run(self, jobs: list[dict], base_resume: str,
            out_dir: Path, max_jobs: int) -> list[dict]:
        out_dir.mkdir(parents=True, exist_ok=True)
        total   = min(len(jobs), max_jobs)
        results = []

        print(f"\nPipeline — {total} jobs → {out_dir}/\n")
        t0 = time.time()

        for i, job in enumerate(jobs[:total], 1):
            results.append(self.process_job(base_resume, job, out_dir, i))

        elapsed = time.time() - t0

        scored  = [r for r in results if isinstance(r["match_score"], int)]
        avg     = sum(r["match_score"] for r in scored) / len(scored) if scored else 0

        cos_scored = [r for r in results if isinstance(r.get("cosine_sim"), float)]
        avg_cos    = (sum(r["cosine_sim"] for r in cos_scored) / len(cos_scored)
                      if cos_scored else None)

        print(f"\n{'═'*60}")
        print(f"Done!  {len(results)} .tex files in '{out_dir}/'")
        print(f"Avg match score   : {avg:.1f}/100  (LLM self-report)")
        if avg_cos is not None:
            print(f"Avg cosine sim    : {avg_cos:.4f}  (semantic overlap, 0–1)")
            # Print a per-job cosine similarity table for quick comparison
            print(f"\n{'─'*60}")
            print(f"  {'#':>3}  {'Cosine':>8}  {'LLM Score':>10}  Title")
            print(f"{'─'*60}")
            for r in results:
                cos = r.get("cosine_sim", "N/A")
                cos_str = f"{cos:.4f}" if isinstance(cos, float) else str(cos)
                ms  = r.get("match_score", "N/A")
                ms_str = f"{ms:>3}/100" if isinstance(ms, int) else f"{'N/A':>6}"
                print(f"  {r['index']:>3}  {cos_str:>8}  {ms_str:>10}  "
                      f"{r['title'][:38]}")
            print(f"{'─'*60}")
        else:
            print(f"Cosine sim        : N/A (sentence-transformers not installed)")
        print(f"Time              : {elapsed:.1f}s")
        print(f"{'═'*60}")


        manifest = out_dir / "_manifest.json"
        manifest.write_text(json.dumps(results, indent=2), encoding="utf-8")
        print(f"Manifest          : {manifest}")

        excel_out = out_dir / "_results.xlsx"
        save_results_excel(results, str(excel_out))
        print()
        return results



def save_results_excel(results: list[dict], out_path: str) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    HDR_BG = "1F3864"  
    HDR_FG = "FFFFFF"
    GREEN_BG = "C6EFCE"   # good score
    YELLOW_BG = "FFEB9C"   
    RED_BG = "FFC7CE"   
    EVEN_BG = "EBF3FB"   
    ODD_BG = "FFFFFF"

    def border():
        s = Side(style="thin", color="C0C0C0")
        return Border(left=s, right=s, top=s, bottom=s)

    def score_fill_llm(val) -> PatternFill:
        try:
            v = int(val)
            if v >= 75: return PatternFill("solid", start_color=GREEN_BG)
            if v >= 50: return PatternFill("solid", start_color=YELLOW_BG)
            return PatternFill("solid", start_color=RED_BG)
        except (ValueError, TypeError):
            return PatternFill("solid", start_color=ODD_BG)

    def score_fill_cos(val) -> PatternFill:
        try:
            v = float(val)
            if v >= 0.55: return PatternFill("solid", start_color=GREEN_BG)
            if v >= 0.35: return PatternFill("solid", start_color=YELLOW_BG)
            return PatternFill("solid", start_color=RED_BG)
        except (ValueError, TypeError):
            return PatternFill("solid", start_color=ODD_BG)

    COLUMNS = [
        ("#",                    6,  "index"),
        ("Job Title",           38,  "title"),
        ("Company",             24,  "company"),
        ("LLM Match Score /100",18,  "match_score"),
        ("Cosine Similarity",   18,  "cosine_sim"),
        ("Keywords Matched",    14,  "kw_count"),
        ("Matched Keywords",    40,  "keywords_str"),
        (".tex File",           48,  "file"),
    ]

    #  Header row 
    hdr_font  = Font(name="Arial", bold=True, color=HDR_FG, size=11)
    hdr_fill  = PatternFill("solid", start_color=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, (label, width, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = hdr_align
        c.border    = border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    #  Data rows 
    data_font  = Font(name="Arial", size=10)
    bold_font  = Font(name="Arial", size=10, bold=True)
    ctr_align  = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(vertical="center", wrap_text=True)

    for ri, r in enumerate(results, 2):
        bg       = EVEN_BG if ri % 2 == 0 else ODD_BG
        row_fill = PatternFill("solid", start_color=bg)

        # Flatten keywords list comma-separated string
        kw_raw = r.get("keywords", [])
        if isinstance(kw_raw, list):
            kw_str   = ", ".join(str(k) for k in kw_raw)
            kw_count = len(kw_raw)
        else:
            kw_str   = str(kw_raw)
            kw_count = 0

        # Build the value dict for this row
        row_data = {
            "index":       r.get("index", ri - 1),
            "title":       r.get("title",       "N/A"),
            "company":     r.get("company",     "N/A"),
            "match_score": r.get("match_score", "N/A"),
            "cosine_sim":  r.get("cosine_sim",  "N/A"),
            "kw_count":    kw_count,
            "keywords_str": kw_str,
            "file":        r.get("file",        "N/A"),
        }

        for ci, (_, _, key) in enumerate(COLUMNS, 1):
            val  = row_data[key]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border()

            if key == "match_score":
                cell.fill      = score_fill_llm(val)
                cell.font      = bold_font
                cell.alignment = ctr_align
            elif key == "cosine_sim":
                if isinstance(val, float):
                    cell.value = round(val, 4)
                cell.fill      = score_fill_cos(val)
                cell.font      = bold_font
                cell.alignment = ctr_align
            elif key in ("index", "kw_count"):
                cell.fill      = row_fill
                cell.font      = data_font
                cell.alignment = ctr_align
            else:
                cell.fill      = row_fill
                cell.font      = data_font
                cell.alignment = wrap_align

        ws.row_dimensions[ri].height = 22

    summary_row = len(results) + 2
    ws.cell(row=summary_row, column=1, value="AVERAGES").font = Font(
        name="Arial", bold=True, size=10)

    scored_llm = [r for r in results if isinstance(r.get("match_score"), int)]
    scored_cos = [r for r in results if isinstance(r.get("cosine_sim"), float)]

    if scored_llm:
        avg_llm = round(sum(r["match_score"] for r in scored_llm) / len(scored_llm), 1)
        c = ws.cell(row=summary_row, column=4, value=avg_llm)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = score_fill_llm(avg_llm)
        c.alignment = Alignment(horizontal="center")
        c.border = border()

    if scored_cos:
        avg_cos = round(sum(r["cosine_sim"] for r in scored_cos) / len(scored_cos), 4)
        c = ws.cell(row=summary_row, column=5, value=avg_cos)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = score_fill_cos(avg_cos)
        c.alignment = Alignment(horizontal="center")
        c.border = border()

    wb.save(out_path)
    print(f"Results Excel     : {out_path}")



COMPILE_HELP = """
  ── COMPILE .tex → PDF ────────────────────────────────────────────
  Install MacTeX (one time):   brew install --cask mactex-no-gui
  Compile one:   pdflatex tailored_resumes/job_001_*.tex
  Compile all:
    for f in tailored_resumes/*.tex; do
      pdflatex -output-directory tailored_resumes "$f"
    done
  ──────────────────────────────────────────────────────────────────
"""

def main():
    parser = argparse.ArgumentParser(
        description="Agentic LaTeX resume tailoring — local Ollama",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            Examples:
              python resume_agent.py --resume my_resume.txt
              python resume_agent.py --resume my_resume.txt --max-jobs 10
              python resume_agent.py --resume my_resume.txt --model mistral
        """),
    )
    parser.add_argument("--resume",   required=True,              help="Path to your resume .txt file")
    parser.add_argument("--excel",    default=DEFAULT_EXCEL,      help="Excel file with job listings")
    parser.add_argument("--max-jobs", type=int, default=DEFAULT_MAX, help="Max number of jobs to process")
    parser.add_argument("--model",    default=DEFAULT_MODEL,      help="Ollama model name (e.g. llama3.2, mistral)")
    parser.add_argument("--out-dir",  default=OUTPUT_DIR,         help="Folder to write .tex output files")
    args = parser.parse_args()

    resume_path = Path(args.resume)
    if not resume_path.exists():
        print(f"Resume not found: {args.resume}"); sys.exit(1)

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"Excel not found: {args.excel}"); sys.exit(1)

    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            base_resume = resume_path.read_text(encoding=enc); break
        except UnicodeDecodeError:
            continue
    else:
        print("Cannot read resume."); sys.exit(1)

    jobs = load_jobs(str(excel_path))
    if not jobs:
        print("No jobs found."); sys.exit(1)

    agent = OllamaResumeAgent(model=args.model)
    agent.run(jobs, base_resume, Path(args.out_dir), args.max_jobs)
    print(COMPILE_HELP)

if __name__ == "__main__":
    main()
