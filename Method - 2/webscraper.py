import re
import sys
import time
import asyncio
import argparse
import textwrap
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from html import unescape
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── AI imports (optional — scraper works without them) ──────────────
try:
    import numpy as np
    import google.generativeai as genai
    from sentence_transformers import SentenceTransformer
    from sklearn.metrics.pairwise import cosine_similarity
    from dotenv import load_dotenv
    import os
    load_dotenv()
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False

# ════════════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════════════

OUTPUT_FILE   = "jobs_output.xlsx"
DEFAULT_PAGES = 10
REQUEST_DELAY = 0.4
TIMEOUT       = 20
MAX_DESC_LEN  = 5000

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html, */*",
}

REMOTIVE_CATEGORIES = [
    "software-dev", "customer-support", "design", "devops", "finance",
    "hr", "management", "marketing", "product", "qa", "sales",
    "data", "writing", "all-others",
]

# ════════════════════════════════════════════════════════════════════
# UTILITIES
# ════════════════════════════════════════════════════════════════════

def html_to_text(html_str: str) -> str:
    if not html_str:
        return "N/A"
    soup = BeautifulSoup(html_str, "html.parser")
    text = soup.get_text(separator="\n")
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    result = "\n".join(lines)
    return result[:MAX_DESC_LEN] if len(result) > MAX_DESC_LEN else result


def clean(val, fallback="N/A") -> str:
    if val is None:
        return fallback
    s = str(val).strip()
    return s if s else fallback


def fmt_date(raw) -> str:
    if not raw:
        return "N/A"
    s = str(raw).strip()
    if s.isdigit():
        try:
            return datetime.fromtimestamp(int(s), tz=timezone.utc).strftime("%Y-%m-%d")
        except Exception:
            pass
    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%a, %d %b %Y %H:%M:%S %z"):
        try:
            return datetime.strptime(s[:25], fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    return s[:20]


def safe_get(url: str, params: dict = None, retries: int = 2) -> requests.Response | None:
    for attempt in range(retries + 1):
        try:
            r = requests.get(url, headers=HEADERS, params=params, timeout=TIMEOUT)
            r.raise_for_status()
            return r
        except requests.exceptions.HTTPError as e:
            print(f"    ✗  HTTP {e.response.status_code} — {url}")
            return None
        except Exception as e:
            if attempt < retries:
                print(f"    ↻  Retry {attempt+1}: {e}")
                time.sleep(2)
            else:
                print(f"    ✗  Failed: {url}  ({e})")
                return None
    return None

# ════════════════════════════════════════════════════════════════════
# SOURCE 1 ▸ REMOTIVE
# ════════════════════════════════════════════════════════════════════

def scrape_remotive(keyword: str = "", category: str = "") -> list[dict]:
    print("\n    Remotive  (remotive.com/api/remote-jobs)")
    params = {}
    if keyword:
        params["search"] = keyword
    if category and category != "all":
        params["category"] = category

    resp = safe_get("https://remotive.com/api/remote-jobs", params=params)
    if not resp:
        return []
    try:
        raw_jobs = resp.json().get("jobs", [])
    except Exception as e:
        print(f"    ✗  JSON parse error: {e}")
        return []

    print(f"    → {len(raw_jobs)} jobs fetched")
    jobs = []
    for j in raw_jobs:
        tags = j.get("tags", [])
        tag_str = ", ".join(tags) if isinstance(tags, list) else clean(tags)
        desc_text = html_to_text(j.get("job_description", ""))
        jobs.append({
            "source":               "Remotive",
            "position_title":       clean(j.get("title")),
            "company":              clean(j.get("company_name")),
            "company_url":          clean(j.get("company_url")),
            "location":             clean(j.get("candidate_required_location") or j.get("location")),
            "work_model":           "Remote",
            "date_posted":          fmt_date(j.get("publication_date")),
            "salary":               clean(j.get("salary")),
            "job_type":             clean(j.get("job_type")),
            "category":             clean(j.get("category")),
            "tags":                 tag_str,
            "qualifications":       _extract_section(desc_text, ["qualif", "requirement", "what we look", "skills"]),
            "responsibilities":     _extract_section(desc_text, ["responsibilit", "what you'll do", "duties"]),
            "benefits":             _extract_section(desc_text, ["benefit", "perks", "what we offer", "we offer"]),
            "experience_required":  _extract_experience(desc_text),
            "education_required":   _extract_education(desc_text),
            "h1b_sponsored":        "Not Specified",
            "is_new_grad":          _is_new_grad(j.get("title", ""), desc_text),
            "apply_url":            clean(j.get("url")),
            "job_description":      desc_text,
            "tailored_resume":      "",   # filled by AI pipeline
            "match_score":          "",   # filled by AI pipeline
        })
    return jobs

# ════════════════════════════════════════════════════════════════════
# SOURCE 2 ▸ ARBEITNOW
# ════════════════════════════════════════════════════════════════════

def scrape_arbeitnow(pages: int = DEFAULT_PAGES, keyword: str = "") -> list[dict]:
    print(f"\n    Arbeitnow  (arbeitnow.com/api/job-board-api)  [up to {pages} pages]")
    base_url = "https://www.arbeitnow.com/api/job-board-api"
    all_raw  = []

    for page_num in range(1, pages + 1):
        resp = safe_get(base_url, params={"page": page_num})
        if not resp:
            break
        try:
            page_jobs = resp.json().get("data", [])
        except Exception as e:
            print(f"    ✗  JSON error on page {page_num}: {e}")
            break
        if not page_jobs:
            print(f"    → No more jobs after page {page_num - 1}")
            break
        all_raw.extend(page_jobs)
        print(f"    Page {page_num:>3}: +{len(page_jobs)} jobs  (total so far: {len(all_raw)})")
        time.sleep(REQUEST_DELAY)

    if keyword:
        kw = keyword.lower()
        all_raw = [j for j in all_raw if kw in (j.get("title","") + j.get("description","")).lower()]
        print(f"    → After keyword filter '{keyword}': {len(all_raw)} jobs")

    jobs = []
    for j in all_raw:
        tags      = j.get("tags", [])
        tag_str   = ", ".join(tags) if isinstance(tags, list) else clean(tags)
        jtypes    = j.get("job_types", [])
        jtype_str = ", ".join(jtypes) if isinstance(jtypes, list) else clean(jtypes)
        desc_text = html_to_text(j.get("description", ""))
        jobs.append({
            "source":               "Arbeitnow",
            "position_title":       clean(j.get("title")),
            "company":              clean(j.get("company_name")),
            "company_url":          "N/A",
            "location":             clean(j.get("location")),
            "work_model":           "Remote" if j.get("remote") else "On-site / Hybrid",
            "date_posted":          fmt_date(j.get("created_at")),
            "salary":               "N/A",
            "job_type":             jtype_str or "N/A",
            "category":             tag_str,
            "tags":                 tag_str,
            "qualifications":       _extract_section(desc_text, ["qualif", "requirement", "what we look", "skills"]),
            "responsibilities":     _extract_section(desc_text, ["responsibilit", "what you'll do", "duties"]),
            "benefits":             _extract_section(desc_text, ["benefit", "perks", "what we offer"]),
            "experience_required":  _extract_experience(desc_text),
            "education_required":   _extract_education(desc_text),
            "h1b_sponsored":        "Not Specified",
            "is_new_grad":          _is_new_grad(j.get("title", ""), desc_text),
            "apply_url":            clean(j.get("url")),
            "job_description":      desc_text,
            "tailored_resume":      "",
            "match_score":          "",
        })
    print(f"    → {len(jobs)} total Arbeitnow jobs")
    return jobs

# ════════════════════════════════════════════════════════════════════
# SOURCE 3 ▸ WE WORK REMOTELY
# ════════════════════════════════════════════════════════════════════

WWR_FEEDS = {
    "programming":      "https://weworkremotely.com/categories/remote-programming-jobs.rss",
    "devops":           "https://weworkremotely.com/categories/remote-devops-sysadmin-jobs.rss",
    "design":           "https://weworkremotely.com/categories/remote-design-jobs.rss",
    "data-science":     "https://weworkremotely.com/categories/remote-data-science-jobs.rss",
    "customer-support": "https://weworkremotely.com/categories/remote-customer-support-jobs.rss",
    "sales-marketing":  "https://weworkremotely.com/categories/remote-sales-and-marketing-jobs.rss",
    "product":          "https://weworkremotely.com/categories/remote-product-jobs.rss",
    "finance":          "https://weworkremotely.com/categories/remote-finance-legal-jobs.rss",
    "writing":          "https://weworkremotely.com/categories/remote-writing-jobs.rss",
}


def scrape_wwr(keyword: str = "", categories: list = None) -> list[dict]:
    feeds = categories if categories else list(WWR_FEEDS.keys())
    print(f"\n    We Work Remotely  (RSS feeds)  [{len(feeds)} categories]")
    all_jobs = []

    for cat in feeds:
        feed_url = WWR_FEEDS.get(cat)
        if not feed_url:
            continue
        resp = safe_get(feed_url)
        if not resp:
            continue
        try:
            root = ET.fromstring(resp.content)
        except ET.ParseError as e:
            print(f"    ✗  XML parse error ({cat}): {e}")
            continue

        cat_jobs = []
        for item in root.findall(".//item"):
            def tag(name):
                el = item.find(name)
                return el.text.strip() if el is not None and el.text else ""

            title_raw = unescape(tag("title"))
            if ":" in title_raw:
                company_part, title_part = title_raw.split(":", 1)
            else:
                company_part, title_part = "N/A", title_raw

            title = title_part.strip()
            if not title or title.lower().startswith("jobs in"):
                continue

            desc_text = html_to_text(unescape(tag("description") or tag("{http://purl.org/rss/1.0/}description")))
            region    = tag("region") or tag("{https://weworkremotely.com}region") or "Worldwide"
            link      = tag("link") or tag("guid")

            if keyword and keyword.lower() not in (title + desc_text).lower():
                continue

            cat_jobs.append({
                "source":               "We Work Remotely",
                "position_title":       title,
                "company":              company_part.strip(),
                "company_url":          "N/A",
                "location":             region or "Worldwide / Remote",
                "work_model":           "Remote",
                "date_posted":          fmt_date(tag("pubDate")),
                "salary":               _extract_salary(desc_text),
                "job_type":             "Full-time",
                "category":             cat.replace("-", " ").title(),
                "tags":                 cat.replace("-", " ").title(),
                "qualifications":       _extract_section(desc_text, ["qualif", "requirement", "skills"]),
                "responsibilities":     _extract_section(desc_text, ["responsibilit", "what you'll do", "duties"]),
                "benefits":             _extract_section(desc_text, ["benefit", "perks", "we offer"]),
                "experience_required":  _extract_experience(desc_text),
                "education_required":   _extract_education(desc_text),
                "h1b_sponsored":        "Not Specified",
                "is_new_grad":          _is_new_grad(title, desc_text),
                "apply_url":            link,
                "job_description":      desc_text,
                "tailored_resume":      "",
                "match_score":          "",
            })

        print(f"    {cat:<20}: {len(cat_jobs)} jobs")
        all_jobs.extend(cat_jobs)
        time.sleep(REQUEST_DELAY)

    print(f"    → {len(all_jobs)} total WWR jobs")
    return all_jobs

# ════════════════════════════════════════════════════════════════════
# FIELD EXTRACTORS
# ════════════════════════════════════════════════════════════════════

def _extract_section(text: str, keywords: list[str]) -> str:
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if any(kw.lower() in line.lower() for kw in keywords):
            section_lines = []
            for j in range(i + 1, min(i + 9, len(lines))):
                l = lines[j].strip()
                if l:
                    section_lines.append(l)
                    if len(section_lines) >= 6:
                        break
            if section_lines:
                return "\n".join(section_lines)
    return "N/A"


def _extract_experience(text: str) -> str:
    patterns = [
        r"(\d+\+?\s*[-–]\s*\d+)\s*years?\s*(?:of\s+)?(?:experience|exp)",
        r"(\d+\+?)\s*years?\s*(?:of\s+)?(?:experience|exp)",
        r"minimum\s+(?:of\s+)?(\d+\+?)\s*years?",
        r"at\s+least\s+(\d+\+?)\s*years?",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(0).strip()[:60]
    return "N/A"


def _extract_education(text: str) -> str:
    edu_map = [
        (r"ph\.?d|doctorate",                              "PhD"),
        (r"master'?s?|m\.?s\.?|mba",                       "Master's Degree"),
        (r"bachelor'?s?|b\.?s\.?|b\.?a\.?|undergraduate",  "Bachelor's Degree"),
        (r"associate'?s?",                                   "Associate's Degree"),
        (r"high\s+school|ged|diploma",                      "High School Diploma"),
        (r"bootcamp|self[-\s]taught|no\s+degree",           "No Degree Required"),
    ]
    for pattern, label in edu_map:
        if re.search(pattern, text, re.IGNORECASE):
            return label
    return "N/A"


def _extract_salary(text: str) -> str:
    patterns = [
        r"\$[\d,]+\s*[-–]\s*\$[\d,]+\s*(?:k|K|per\s+year|\/yr|\/year|annually)?",
        r"\$[\d,]+[kK]?\s*[-–]\s*\$[\d,]+[kK]?",
        r"[\d,]+\s*[-–]\s*[\d,]+\s*(?:USD|EUR|GBP)",
        r"salary[:\s]+\$[\d,]+",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return m.group(0).strip()[:60]
    return "N/A"


def _is_new_grad(title: str, desc: str) -> str:
    combined = (title + " " + desc).lower()
    if re.search(r"new\s+grad|entry[-\s]level|junior|jr\.?|graduate|fresh(er|man)?|0[-–]2\s*years?", combined):
        return "Yes"
    if re.search(r"senior|staff|principal|lead|manager|director|vp\b|head\s+of|10\+\s*years?", combined):
        return "No"
    return "Not Specified"

# ════════════════════════════════════════════════════════════════════
# AI PIPELINE  (Gemini + sentence-transformers)
# ════════════════════════════════════════════════════════════════════

class ResumeAgent:
    """
    Single-pass agentic resume tailor using Gemini.

    Each job gets ONE Gemini call that:
      1. Extracts key requirements from the JD
      2. Scores the base resume (0–100)
      3. Returns a tailored resume

    Embedding similarity is computed locally — no extra LLM call needed.
    """

    SYSTEM_PROMPT = """You are an expert resume coach and technical recruiter.
Given a job description and a base resume, you must return a JSON object with exactly these keys:
{
  "key_requirements": ["skill1", "skill2", ...],   // top 5-8 must-haves from the JD
  "match_score": 72,                               // integer 0-100, how well base resume fits
  "tailored_resume": "..."                         // rewritten resume optimised for this JD
}

Rules for tailored_resume:
- Keep the same overall structure as the base resume
- Rephrase bullets to mirror language from the JD
- Add any relevant skills the candidate likely has but didn't explicitly list
- Do NOT invent jobs, companies, or degrees
- Keep it concise — under 400 words

Respond with ONLY valid JSON. No markdown, no preamble."""

    def __init__(self, api_key: str):
        genai.configure(api_key=api_key)
        self.model       = genai.GenerativeModel("gemini-1.5-flash")
        self.embed_model = SentenceTransformer("all-MiniLM-L6-v2")

    def _embedding_similarity(self, text_a: str, text_b: str) -> float:
        embs = self.embed_model.encode([text_a, text_b])
        return float(cosine_similarity([embs[0]], [embs[1]])[0][0])

    def process_job(self, base_resume: str, job: dict) -> dict:
        """
        Run the single-pass pipeline for one job.
        Returns the job dict updated with tailored_resume and match_score.
        """
        jd = job.get("job_description", "")
        if not jd or jd == "N/A":
            job["tailored_resume"] = "N/A"
            job["match_score"]     = "N/A"
            return job

        prompt = (
            f"JOB DESCRIPTION:\n{jd[:3000]}\n\n"
            f"BASE RESUME:\n{base_resume}"
        )

        try:
            response = self.model.generate_content(
                [{"role": "user", "parts": [self.SYSTEM_PROMPT + "\n\n" + prompt]}],
                generation_config={"temperature": 0.3, "max_output_tokens": 1024},
            )
            import json
            raw = response.text.strip().lstrip("```json").rstrip("```").strip()
            data = json.loads(raw)

            tailored = data.get("tailored_resume", base_resume)
            score    = int(data.get("match_score", 50))

            # Boost score with local embedding similarity (no extra API call)
            emb_sim  = self._embedding_similarity(tailored, jd)
            final_score = round(0.7 * score + 0.3 * emb_sim * 100)

            job["tailored_resume"] = tailored
            job["match_score"]     = final_score

        except Exception as e:
            print(f"      ⚠  AI error for '{job.get('position_title','?')}': {e}")
            job["tailored_resume"] = base_resume
            job["match_score"]     = "Error"

        return job

    def run_pipeline(
        self,
        jobs: list[dict],
        base_resume: str,
        max_jobs: int = 20,
        delay: float = 1.0,
    ) -> list[dict]:
        """
        Process up to `max_jobs` jobs sequentially (polite to the API).
        Jobs beyond max_jobs are left with empty tailored_resume.
        """
        target = [j for j in jobs if j.get("job_description", "N/A") != "N/A"][:max_jobs]
        total  = len(target)
        print(f"\n  🤖  AI pipeline: tailoring resume for {total} jobs …")

        start = time.time()
        for i, job in enumerate(target, 1):
            print(f"      [{i:>3}/{total}]  {job.get('position_title','?')[:50]}")
            self.process_job(base_resume, job)
            time.sleep(delay)

        elapsed = time.time() - start
        scored  = [j for j in target if isinstance(j.get("match_score"), int)]
        if scored:
            avg = sum(j["match_score"] for j in scored) / len(scored)
            print(f"\n  📊  Average match score : {avg:.1f}/100")
        print(f"  ⏱   Time elapsed        : {elapsed:.1f}s  ({elapsed/total:.1f}s/job)")

        return jobs   # full list — un-processed jobs have empty fields

# ════════════════════════════════════════════════════════════════════
# EXCEL WRITER
# ════════════════════════════════════════════════════════════════════

COLUMNS = [
    ("source",              "Source",                 16),
    ("position_title",      "Position Title",         32),
    ("company",             "Company",                22),
    ("location",            "Location",               22),
    ("work_model",          "Work Model",             14),
    ("date_posted",         "Date Posted",            13),
    ("salary",              "Salary",                 22),
    ("job_type",            "Job Type",               16),
    ("category",            "Category",               20),
    ("tags",                "Tags / Skills",          28),
    ("experience_required", "Experience Required",    22),
    ("education_required",  "Education Required",     20),
    ("qualifications",      "Qualifications",         38),
    ("responsibilities",    "Responsibilities",       38),
    ("benefits",            "Benefits",               28),
    ("h1b_sponsored",       "H1B Sponsored",          14),
    ("is_new_grad",         "Is New Grad",            13),
    ("match_score",         "AI Match Score",         14),   # NEW
    ("apply_url",           "Apply Link",             42),
    ("tailored_resume",     "Tailored Resume (AI)",   55),   # NEW
    ("job_description",     "Full Job Description",   55),
    ("company_url",         "Company URL",            35),
]

HDR_BG    = "1F3864"
HDR_FG    = "FFFFFF"
EVEN_BG   = "EBF3FB"
ODD_BG    = "FFFFFF"
GREEN_BG  = "C6EFCE"
RED_BG    = "FFC7CE"
YELLOW_BG = "FFEB9C"
BLUE_BG   = "BDD7EE"

SOURCE_COLORS = {
    "Remotive":         "D9EAD3",
    "Arbeitnow":        "FCE5CD",
    "We Work Remotely": "CFE2F3",
}


def _border():
    s = Side(style="thin", color="C0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)


def _score_fill(val) -> PatternFill:
    """Colour-code the AI match score cell."""
    try:
        v = int(val)
        if v >= 75:
            return PatternFill("solid", start_color=GREEN_BG)
        if v >= 50:
            return PatternFill("solid", start_color=YELLOW_BG)
        return PatternFill("solid", start_color=RED_BG)
    except (ValueError, TypeError):
        return PatternFill("solid", start_color=ODD_BG)


def _write_jobs_sheet(wb: openpyxl.Workbook, jobs: list[dict]):
    ws = wb.active
    ws.title = "Jobs"

    hdr_font  = Font(name="Arial", bold=True, color=HDR_FG, size=11)
    hdr_fill  = PatternFill("solid", start_color=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr       = _border()

    for ci, (_, label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="1155CC", underline="single")
    top_wrap  = Alignment(vertical="top", wrap_text=True)
    top_ctr   = Alignment(horizontal="center", vertical="top", wrap_text=True)

    for ri, job in enumerate(jobs, 2):
        bg       = EVEN_BG if ri % 2 == 0 else ODD_BG
        row_fill = PatternFill("solid", start_color=bg)

        for ci, (key, _, _) in enumerate(COLUMNS, 1):
            val  = job.get(key, "N/A") or "N/A"
            cell = ws.cell(row=ri, column=ci, value=str(val))
            cell.border = bdr

            is_link   = key in ("apply_url", "company_url") and str(val).startswith("http")
            is_center = key in ("source", "work_model", "date_posted", "job_type",
                                "h1b_sponsored", "is_new_grad", "experience_required",
                                "education_required", "match_score")
            is_flag   = key in ("h1b_sponsored", "is_new_grad")
            is_score  = key == "match_score"

            if key == "source":
                cell.fill = PatternFill("solid", start_color=SOURCE_COLORS.get(str(val), ODD_BG))
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.alignment = top_ctr
            elif is_score:
                cell.fill = _score_fill(val)
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.alignment = top_ctr
            elif is_link:
                cell.fill = row_fill
                cell.font = link_font
                cell.hyperlink = str(val)
                cell.alignment = top_wrap
            elif is_flag:
                v = str(val).lower()
                cell.fill = PatternFill("solid", start_color=GREEN_BG if v == "yes" else RED_BG if v == "no" else YELLOW_BG)
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.alignment = top_ctr
            elif is_center:
                cell.fill = row_fill
                cell.font = data_font
                cell.alignment = top_ctr
            else:
                cell.fill = row_fill
                cell.font = data_font
                cell.alignment = top_wrap

        ws.row_dimensions[ri].height = 55


def _write_summary_sheet(wb: openpyxl.Workbook, jobs: list[dict], args):
    from collections import Counter
    ws = wb.create_sheet("Summary")

    ws.merge_cells("A1:C1")
    ws["A1"].value     = "Job Scrape Summary Report"
    ws["A1"].font      = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill      = PatternFill("solid", start_color=HDR_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    def row(r, label, val, bold_val=False):
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=val)
        c1.font = Font(name="Arial", bold=True, size=11)
        c2.font = Font(name="Arial", bold=bold_val, size=11)
        for c in (c1, c2):
            c.border = _border()
            c.alignment = Alignment(vertical="center", wrap_text=True)

    row(2, "Scraped At",       datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
    row(3, "Total Jobs",       len(jobs), bold_val=True)
    row(4, "Keyword Filter",   args.keyword or "None")
    row(5, "Sources Used",     ", ".join(args.sources))
    row(6, "AI Processed",     sum(1 for j in jobs if j.get("match_score") not in ("", "N/A", "Error", None)))

    # Source breakdown
    ws.cell(row=8, column=1, value="Jobs by Source").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(row=8, column=1).fill = PatternFill("solid", start_color="4472C4")
    ws.cell(row=8, column=2, value="Count").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(row=8, column=2).fill = PatternFill("solid", start_color="4472C4")

    source_counts = Counter(j.get("source", "?") for j in jobs)
    for i, (src, cnt) in enumerate(source_counts.items(), start=9):
        ws.cell(row=i, column=1, value=src).font = Font(name="Arial", size=11)
        ws.cell(row=i, column=2, value=cnt).font  = Font(name="Arial", size=11)
        src_bg = SOURCE_COLORS.get(src, ODD_BG)
        ws.cell(row=i, column=1).fill = PatternFill("solid", start_color=src_bg)
        ws.cell(row=i, column=2).fill = PatternFill("solid", start_color=src_bg)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20


def save_excel(jobs: list[dict], args, filepath: str):
    print(f"\n    Writing {len(jobs)} jobs to Excel …")
    wb = openpyxl.Workbook()
    _write_jobs_sheet(wb, jobs)
    _write_summary_sheet(wb, jobs, args)
    wb.save(filepath)
    print(f"    Saved → {filepath}")

# ════════════════════════════════════════════════════════════════════
# CLI
# ════════════════════════════════════════════════════════════════════


def main():
    parser = argparse.ArgumentParser(
        description="Scrape live job listings and optionally tailor your resume with AI",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            Examples:
              python webscraper.py
              python webscraper.py --keyword "data analyst"
              python webscraper.py --sources remotive wwr
              python webscraper.py --keyword "python" --pages 5 --output python_jobs.xlsx
              python webscraper.py --resume my_resume.txt --ai-jobs 20
              python webscraper.py --sources arbeitnow --pages 20

            Remotive categories:
              software-dev, customer-support, design, devops, finance,
              hr, management, marketing, product, qa, sales, data, writing
        """),
    )
    parser.add_argument("--sources",  nargs="+", choices=["remotive","arbeitnow","wwr"],
                        default=["remotive","arbeitnow","wwr"])
    parser.add_argument("--keyword",  default="",            help="Keyword filter")
    parser.add_argument("--category", default="",            help="Remotive category filter")
    parser.add_argument("--pages",    type=int, default=DEFAULT_PAGES, help="Arbeitnow pages")
    parser.add_argument("--output",   default=OUTPUT_FILE,   help="Output Excel filename")
    parser.add_argument("--resume",   default="",
                        help="Path to your plain-text resume (enables AI tailoring)")
    parser.add_argument("--ai-jobs",  type=int, default=20,
                        help="Max jobs to process with AI (default: 20)")
    args = parser.parse_args()

    print(BANNER)
    print(f"  Sources  : {', '.join(args.sources)}")
    print(f"  Keyword  : {args.keyword or '(none)'}")
    print(f"  Output   : {args.output}")
    if args.resume:
        print(f"  Resume   : {args.resume}  (AI tailoring ON, up to {args.ai_jobs} jobs)")
    print()

    # ── Scrape ──────────────────────────────────────────────────────
    all_jobs: list[dict] = []
    try:
        if "remotive"  in args.sources:
            all_jobs.extend(scrape_remotive(keyword=args.keyword, category=args.category))
        if "arbeitnow" in args.sources:
            all_jobs.extend(scrape_arbeitnow(pages=args.pages, keyword=args.keyword))
        if "wwr"       in args.sources:
            all_jobs.extend(scrape_wwr(keyword=args.keyword))
    except KeyboardInterrupt:
        print("\n\n  Interrupted — saving what we have …")

    if not all_jobs:
        print("\n  No jobs extracted. Check your internet connection.")
        sys.exit(1)

    # ── Deduplicate ─────────────────────────────────────────────────
    seen, unique_jobs = set(), []
    for j in all_jobs:
        key = (j.get("position_title","").lower().strip(),
               j.get("company","").lower().strip())
        if key not in seen:
            seen.add(key)
            unique_jobs.append(j)
    removed = len(all_jobs) - len(unique_jobs)
    if removed:
        print(f"\nRemoved {removed} duplicate(s)  →  {len(unique_jobs)} unique jobs")

    # ── AI pipeline (optional) ──────────────────────────────────────
    if args.resume:
        resume_path = Path(args.resume)
        if not resume_path.exists():
            print(f"\n  Resume file not found: {args.resume} — skipping AI step.")
        elif not AI_AVAILABLE:
            print("\n  AI deps missing. Run: pip install google-generativeai sentence-transformers scikit-learn numpy python-dotenv")
        else:
            api_key = os.getenv("GOOGLE_API_KEY", "")
            if not api_key:
                print("\n  GOOGLE_API_KEY not set in .env — skipping AI step.")
            else:
                base_resume = resume_path.read_text(encoding="utf-8")
                agent = ResumeAgent(api_key=api_key)
                unique_jobs = agent.run_pipeline(
                    unique_jobs, base_resume, max_jobs=args.ai_jobs
                )

    # ── Save ────────────────────────────────────────────────────────
    save_excel(unique_jobs, args, args.output)

    print(f"\n{'═'*56}")
    print(f"    Done!  {len(unique_jobs)} jobs saved to '{args.output}'")
    print(f"    Open the file — it has 2 sheets: Jobs + Summary")
    print(f"{'═'*56}\n")


if __name__ == "__main__":
    main()